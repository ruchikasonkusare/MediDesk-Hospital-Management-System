"""
MediDesk Export Module — Excel (openpyxl) + PDF (reportlab)
"""
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from datetime import date


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _excel_response(filename):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _pdf_response(filename):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# Excel — Patient List
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def export_patients_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('Run: pip install openpyxl', status=500)

    from core.models import Patient

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Patients'

    TEAL   = '0B7A6E'
    WHITE  = 'FFFFFF'
    ALT    = 'F0F9F8'
    BORDER = 'DDDDDD'

    hdr_font  = Font(bold=True, color=WHITE, size=11)
    hdr_fill  = PatternFill('solid', fgColor=TEAL)
    alt_fill  = PatternFill('solid', fgColor=ALT)
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_side = Side(style='thin', color=BORDER)
    thin      = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Title
    ws.merge_cells('A1:M1')
    t = ws['A1']
    t.value = f'Patient Registry  |  MediDesk  |  {date.today().strftime("%d %b %Y")}'
    t.font  = Font(bold=True, size=13, color=TEAL)
    t.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    # Headers
    headers = [
        'Patient ID','Name','Age','Gender','Blood Group','Phone',
        'Email','City','Chronic Conditions','Allergies',
        'Insurance Provider','Insurance No.','Registered On',
    ]
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = thin

    ws.row_dimensions[2].height = 20

    # Data
    for row_idx, patient in enumerate(Patient.objects.all().order_by('name'), 3):
        row = [
            patient.patient_id, patient.name, patient.age,
            patient.get_gender_display(), patient.blood_group or '—',
            patient.phone, patient.email or '—',
            getattr(patient, 'city', '') or '—',
            patient.chronic_conditions or '—',
            patient.allergies or '—',
            patient.insurance_provider or '—',
            patient.insurance_number or '—',
            patient.registered_at.strftime('%d %b %Y'),
        ]
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border    = thin
            if fill:
                cell.fill = fill

    # Column widths
    widths = [12, 22, 6, 8, 10, 13, 24, 14, 28, 22, 20, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Freeze header rows
    ws.freeze_panes = 'A3'

    response = _excel_response(f'patients_{date.today()}.xlsx')
    wb.save(response)
    return response


# ─────────────────────────────────────────────────────────────────────────────
# Excel — Billing Summary
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def export_billing_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('Run: pip install openpyxl', status=500)

    from core.models import Bill
    from django.db.models import Sum

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Billing'

    BLUE      = '1A56B0'
    WHITE     = 'FFFFFF'
    ALT       = 'EEF4FF'
    BORDER    = 'DDDDDD'
    hdr_font  = Font(bold=True, color=WHITE, size=11)
    hdr_fill  = PatternFill('solid', fgColor=BLUE)
    alt_fill  = PatternFill('solid', fgColor=ALT)
    thin_side = Side(style='thin', color=BORDER)
    thin      = Border(left=thin_side, right=thin_side,
                       top=thin_side, bottom=thin_side)
    center    = Alignment(horizontal='center', vertical='center')

    # Filters from query params
    status_filter = request.GET.get('status', '')
    qs = Bill.objects.select_related('patient').all()
    if status_filter:
        qs = qs.filter(payment_status=status_filter)

    # Title
    ws.merge_cells('A1:K1')
    t = ws['A1']
    status_label = f' — Status: {status_filter.title()}' if status_filter else ''
    t.value = f'Billing Summary{status_label}  |  MediDesk  |  {date.today().strftime("%d %b %Y")}'
    t.font  = Font(bold=True, size=13, color=BLUE)
    ws.row_dimensions[1].height = 28

    headers = [
        'Bill No.', 'Patient', 'Bill Date',
        'Consultation', 'Procedure', 'Lab',
        'Gross (₹)', 'Discount (₹)', 'Net (₹)',
        'Paid (₹)', 'Status',
    ]
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = thin
    ws.row_dimensions[2].height = 20

    for ri, bill in enumerate(qs.order_by('-bill_date'), 3):
        fill = alt_fill if ri % 2 == 0 else None
        row  = [
            bill.bill_number, bill.patient.name,
            bill.bill_date.strftime('%d %b %Y'),
            float(bill.consultation_charge),
            float(bill.procedure_charge),
            float(bill.lab_charge),
            float(bill.gross_amount),
            float(bill.discount_amount),
            float(bill.net_amount),
            float(bill.amount_paid),
            bill.get_payment_status_display(),
        ]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = thin
            cell.alignment = Alignment(vertical='center')
            if isinstance(val, float):
                cell.number_format = '#,##0.00'
                cell.alignment     = Alignment(horizontal='right', vertical='center')
            if fill:
                cell.fill = fill

    # Totals row
    total_row = ri + 1
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    for col, field in [(7,'gross_amount'),(8,'discount_amount'),
                       (9,'net_amount'),(10,'amount_paid')]:
        total = float(qs.aggregate(s=Sum(field))['s'] or 0)
        cell = ws.cell(row=total_row, column=col, value=total)
        cell.font          = Font(bold=True)
        cell.number_format = '#,##0.00'
        cell.alignment     = Alignment(horizontal='right')

    widths = [14, 22, 12, 13, 12, 10, 12, 13, 12, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    response = _excel_response(f'billing_{date.today()}.xlsx')
    wb.save(response)
    return response


# ─────────────────────────────────────────────────────────────────────────────
# PDF — Prescription
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def export_prescription_pdf(request, pk):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table,
            TableStyle, HRFlowable,
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    except ImportError:
        return HttpResponse('Run: pip install reportlab', status=500)

    from core.models import Prescription

    rx = get_object_or_404(Prescription.objects.prefetch_related('items'), pk=pk)

    response = _pdf_response(f'prescription_{rx.pk}.pdf')
    doc = SimpleDocTemplate(response, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    TEAL  = colors.HexColor('#0B7A6E')
    DARK  = colors.HexColor('#0d1117')
    GRAY  = colors.HexColor('#7a8699')
    LGRAY = colors.HexColor('#f7f8fa')

    styles = getSampleStyleSheet()
    h1  = ParagraphStyle('h1',  fontSize=22, textColor=TEAL,
                         fontName='Helvetica-Bold', spaceAfter=2)
    h2  = ParagraphStyle('h2',  fontSize=11, textColor=DARK,
                         fontName='Helvetica-Bold', spaceAfter=4)
    sub = ParagraphStyle('sub', fontSize=9,  textColor=GRAY, spaceAfter=2)
    bod = ParagraphStyle('bod', fontSize=10, textColor=DARK, spaceAfter=4)
    sml = ParagraphStyle('sml', fontSize=8,  textColor=GRAY)

    story = []

    # Header
    story.append(Paragraph('MediDesk', h1))
    story.append(Paragraph('Healthcare Management System', sub))
    story.append(HRFlowable(width='100%', thickness=1.5, color=TEAL))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph('PRESCRIPTION', ParagraphStyle(
        'rx_title', fontSize=14, textColor=DARK,
        fontName='Helvetica-Bold', alignment=TA_CENTER)))
    story.append(Spacer(1, 0.4*cm))

    # Patient + Doctor info
    info_data = [
        ['Patient', rx.patient.name,
         'Doctor', f'Dr. {rx.doctor.name}'],
        ['Patient ID', rx.patient.patient_id,
         'Specialization', rx.doctor.get_specialization_display()],
        ['Age / Gender',
         f'{rx.patient.age} yrs / {rx.patient.get_gender_display()}',
         'Date', rx.prescribed_date.strftime('%d %b %Y')],
        ['Blood Group', rx.patient.blood_group or '—',
         'Valid Until',
         rx.valid_until.strftime('%d %b %Y') if rx.valid_until else '—'],
    ]
    info_table = Table(info_data, colWidths=[3*cm, 6*cm, 3.5*cm, 5*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME',  (0,0),(-1,-1), 'Helvetica'),
        ('FONTNAME',  (0,0),(0,-1),  'Helvetica-Bold'),
        ('FONTNAME',  (2,0),(2,-1),  'Helvetica-Bold'),
        ('FONTSIZE',  (0,0),(-1,-1), 9),
        ('TEXTCOLOR', (0,0),(0,-1),  GRAY),
        ('TEXTCOLOR', (2,0),(2,-1),  GRAY),
        ('TEXTCOLOR', (1,0),(1,-1),  DARK),
        ('TEXTCOLOR', (3,0),(3,-1),  DARK),
        ('ROWBACKGROUNDS', (0,0),(-1,-1), [LGRAY, colors.white]),
        ('TOPPADDING',    (0,0),(-1,-1), 5),
        ('BOTTOMPADDING', (0,0),(-1,-1), 5),
        ('LEFTPADDING',   (0,0),(-1,-1), 6),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.4*cm))

    # Alerts
    if rx.patient.allergies:
        story.append(Paragraph(
            f'⚠ ALLERGY ALERT: {rx.patient.allergies}',
            ParagraphStyle('alert', fontSize=9, textColor=colors.HexColor('#C0392B'),
                           fontName='Helvetica-Bold',
                           backColor=colors.HexColor('#FDECEA'),
                           leftPadding=8, rightPadding=8,
                           topPadding=4, bottomPadding=4)
        ))
        story.append(Spacer(1, 0.3*cm))

    # Medicines
    story.append(Paragraph('Medicines', h2))
    med_headers = ['#', 'Medicine', 'Dosage', 'Frequency', 'Timing', 'Days', 'Qty']
    med_rows = [med_headers]
    for idx, item in enumerate(rx.items.all(), 1):
        med_rows.append([
            str(idx),
            f'{item.medicine_name}\n{item.generic_name or ""}',
            item.dosage,
            item.get_frequency_display(),
            item.get_timing_display() if item.timing else '—',
            str(item.duration_days),
            str(item.quantity),
        ])

    med_table = Table(med_rows, colWidths=[0.6*cm, 5.5*cm, 2.2*cm, 3*cm, 2.5*cm, 1.2*cm, 1.2*cm])
    med_table.setStyle(TableStyle([
        ('BACKGROUND',   (0,0), (-1,0),  TEAL),
        ('TEXTCOLOR',    (0,0), (-1,0),  colors.white),
        ('FONTNAME',     (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',     (0,0), (-1,-1), 8.5),
        ('FONTNAME',     (0,1), (-1,-1), 'Helvetica'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.white, LGRAY]),
        ('GRID',         (0,0), (-1,-1), 0.3, colors.HexColor('#e2e6ed')),
        ('ALIGN',        (0,0), (-1,-1), 'LEFT'),
        ('ALIGN',        (5,0), (6,-1),  'CENTER'),
        ('TOPPADDING',   (0,0), (-1,-1), 5),
        ('BOTTOMPADDING',(0,0), (-1,-1), 5),
        ('LEFTPADDING',  (0,0), (-1,-1), 5),
    ]))
    story.append(med_table)
    story.append(Spacer(1, 0.4*cm))

    # Notes
    if rx.notes:
        story.append(Paragraph('Instructions', h2))
        story.append(Paragraph(rx.notes, bod))
        story.append(Spacer(1, 0.3*cm))

    # Footer / Signature
    story.append(HRFlowable(width='100%', thickness=0.5, color=GRAY))
    story.append(Spacer(1, 0.3*cm))
    footer_data = [
        [Paragraph('Doctor Signature', sml),
         Paragraph('', sml),
         Paragraph(f'Printed on: {date.today().strftime("%d %b %Y")}', sml)],
        [Paragraph(f'Dr. {rx.doctor.name}', bod),
         '',
         Paragraph('MediDesk Healthcare System', sml)],
    ]
    footer_table = Table(footer_data, colWidths=[8*cm, 3*cm, 6.5*cm])
    footer_table.setStyle(TableStyle([
        ('ALIGN', (2,0), (2,-1), 'RIGHT'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
    ]))
    story.append(footer_table)

    doc.build(story)
    return response


# ─────────────────────────────────────────────────────────────────────────────
# PDF — Bill / Invoice
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def export_bill_pdf(request, pk):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table,
            TableStyle, HRFlowable,
        )
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    except ImportError:
        return HttpResponse('Run: pip install reportlab', status=500)

    from core.models import Bill

    bill = get_object_or_404(Bill.objects.select_related('patient')
                             .prefetch_related('payments'), pk=pk)

    response = _pdf_response(f'bill_{bill.bill_number}.pdf')
    doc = SimpleDocTemplate(response, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    BLUE  = colors.HexColor('#1A56B0')
    DARK  = colors.HexColor('#0d1117')
    GRAY  = colors.HexColor('#7a8699')
    LGRAY = colors.HexColor('#f7f8fa')
    GREEN = colors.HexColor('#1a7a3e')
    RED   = colors.HexColor('#C0392B')
    AMBER = colors.HexColor('#e87c1e')

    styles = getSampleStyleSheet()
    h1  = ParagraphStyle('h1',  fontSize=22, textColor=BLUE,
                         fontName='Helvetica-Bold')
    sub = ParagraphStyle('sub', fontSize=9,  textColor=GRAY)
    h2  = ParagraphStyle('h2',  fontSize=11, textColor=DARK,
                         fontName='Helvetica-Bold', spaceAfter=6)
    sml = ParagraphStyle('sml', fontSize=8, textColor=GRAY)

    story = []

    # Header
    story.append(Paragraph('MediDesk', h1))
    story.append(Paragraph('Tax Invoice / Receipt', sub))
    story.append(HRFlowable(width='100%', thickness=1.5, color=BLUE))
    story.append(Spacer(1, 0.4*cm))

    # Bill meta
    status_color = GREEN if bill.payment_status == 'paid' else \
                   AMBER if bill.payment_status == 'partial' else RED
    story.append(Paragraph(
        f'<font color="#{bill.payment_status.upper()}">Bill No: {bill.bill_number}</font>',
        ParagraphStyle('bn', fontSize=13, fontName='Helvetica-Bold', textColor=DARK)))
    story.append(Spacer(1, 0.2*cm))

    meta = [
        ['Patient', bill.patient.name,
         'Bill Date', bill.bill_date.strftime('%d %b %Y')],
        ['Patient ID', bill.patient.patient_id,
         'Status',
         bill.get_payment_status_display()],
        ['Phone', bill.patient.phone,
         'Payment Method',
         bill.payment_method.replace('_',' ').title() if bill.payment_method else '—'],
    ]
    meta_table = Table(meta, colWidths=[3*cm, 6*cm, 3.5*cm, 5*cm])
    meta_table.setStyle(TableStyle([
        ('FONTNAME',  (0,0),(0,-1), 'Helvetica-Bold'),
        ('FONTNAME',  (2,0),(2,-1), 'Helvetica-Bold'),
        ('FONTNAME',  (1,0),(-1,-1),'Helvetica'),
        ('FONTSIZE',  (0,0),(-1,-1), 9),
        ('TEXTCOLOR', (0,0),(0,-1),  GRAY),
        ('TEXTCOLOR', (2,0),(2,-1),  GRAY),
        ('ROWBACKGROUNDS',(0,0),(-1,-1),[LGRAY, colors.white]),
        ('TOPPADDING',    (0,0),(-1,-1), 5),
        ('BOTTOMPADDING', (0,0),(-1,-1), 5),
        ('LEFTPADDING',   (0,0),(-1,-1), 6),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 0.5*cm))

    # Charges
    story.append(Paragraph('Charges Breakdown', h2))
    charge_rows = [['Description', 'Amount (₹)']]
    charges = [
        ('Consultation Fee',  bill.consultation_charge),
        ('Procedure Charges', bill.procedure_charge),
        ('Medicine Charges',  bill.medicine_charge),
        ('Lab / Investigation', bill.lab_charge),
        ('Room Charges',      bill.room_charge),
        (bill.other_charge_label or 'Other', bill.other_charge),
    ]
    for label, amt in charges:
        if float(amt) > 0:
            charge_rows.append([label, f'{float(amt):,.2f}'])

    charge_table = Table(charge_rows, colWidths=[13*cm, 4.5*cm])
    charge_table.setStyle(TableStyle([
        ('BACKGROUND',   (0,0),(-1,0),  BLUE),
        ('TEXTCOLOR',    (0,0),(-1,0),  colors.white),
        ('FONTNAME',     (0,0),(-1,0),  'Helvetica-Bold'),
        ('FONTNAME',     (0,1),(-1,-1), 'Helvetica'),
        ('FONTSIZE',     (0,0),(-1,-1), 9.5),
        ('ALIGN',        (1,0),(1,-1),  'RIGHT'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, LGRAY]),
        ('GRID',         (0,0),(-1,-1), 0.3, colors.HexColor('#e2e6ed')),
        ('TOPPADDING',   (0,0),(-1,-1), 5),
        ('BOTTOMPADDING',(0,0),(-1,-1), 5),
        ('LEFTPADDING',  (0,0),(-1,-1), 8),
    ]))
    story.append(charge_table)
    story.append(Spacer(1, 0.2*cm))

    # Totals
    totals = [
        ['Gross Total', f'₹ {float(bill.gross_amount):,.2f}'],
    ]
    if float(bill.discount_amount) > 0:
        totals.append([f'Discount ({bill.discount_percent}%)',
                       f'- ₹ {float(bill.discount_amount):,.2f}'])
    if float(getattr(bill,'tax_percent',0) or 0) > 0:
        totals.append([f'Tax ({bill.tax_percent}%)',
                       f'+ ₹ {float(bill.net_amount - bill.gross_amount + bill.discount_amount):,.2f}'])
    totals.append(['NET AMOUNT', f'₹ {float(bill.net_amount):,.2f}'])
    totals.append(['Amount Paid', f'₹ {float(bill.amount_paid):,.2f}'])
    totals.append(['Balance Due', f'₹ {float(bill.balance_due):,.2f}'])

    totals_table = Table(totals, colWidths=[13*cm, 4.5*cm])
    totals_style = [
        ('FONTNAME',  (0,0),(-1,-1), 'Helvetica'),
        ('FONTSIZE',  (0,0),(-1,-1), 9.5),
        ('ALIGN',     (1,0),(1,-1),  'RIGHT'),
        ('TOPPADDING',(0,0),(-1,-1), 4),
        ('BOTTOMPADDING',(0,0),(-1,-1), 4),
        ('LEFTPADDING',(0,0),(-1,-1), 8),
    ]
    # Highlight NET row
    net_idx = next(i for i,r in enumerate(totals) if r[0]=='NET AMOUNT')
    totals_style += [
        ('FONTNAME',    (0,net_idx),(1,net_idx), 'Helvetica-Bold'),
        ('BACKGROUND',  (0,net_idx),(1,net_idx), LGRAY),
        ('FONTSIZE',    (0,net_idx),(1,net_idx), 11),
        ('TEXTCOLOR',   (1,net_idx),(1,net_idx), BLUE),
    ]
    # Highlight Balance Due
    bal_idx = len(totals) - 1
    bal_color = GREEN if float(bill.balance_due) == 0 else RED
    totals_style += [
        ('FONTNAME',   (0,bal_idx),(1,bal_idx), 'Helvetica-Bold'),
        ('TEXTCOLOR',  (1,bal_idx),(1,bal_idx), bal_color),
    ]
    totals_table.setStyle(TableStyle(totals_style))
    story.append(totals_table)
    story.append(Spacer(1, 0.5*cm))

    # Footer
    story.append(HRFlowable(width='100%', thickness=0.5,
                             color=colors.HexColor('#e2e6ed')))
    story.append(Spacer(1, 0.3*cm))
    footer = Table([
        [Paragraph('Thank you for choosing MediDesk', sml),
         Paragraph(f'Generated: {date.today().strftime("%d %b %Y")}', sml)],
    ], colWidths=[10*cm, 7.5*cm])
    footer.setStyle(TableStyle([('ALIGN',(1,0),(1,-1),'RIGHT')]))
    story.append(footer)

    doc.build(story)
    return response
